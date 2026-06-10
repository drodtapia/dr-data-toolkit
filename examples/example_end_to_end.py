from pathlib import Path

import pandas as pd
from openpyxl import Workbook

from dr_datakit import (
    load_file,
    clean_df,
    validate_not_empty,
    validate_required_columns,
    validate_not_null,
    validate_unique,
    validate_non_negative,
    validate_allowed_values,
    export_csv,
    export_parquet,
    create_report,
)


# ======================================================
# 1. PROJECT PATHS
# ======================================================

BASE_DIR = Path(".")
DATA_DIR = BASE_DIR / "data"
RAW_DIR = DATA_DIR / "raw"
PROCESSED_DIR = DATA_DIR / "processed"
OUTPUT_DIR = BASE_DIR / "output"
REPORTS_DIR = OUTPUT_DIR / "reports"
TEMPLATES_DIR = BASE_DIR / "templates"

for folder in [RAW_DIR, PROCESSED_DIR, REPORTS_DIR, TEMPLATES_DIR]:
    folder.mkdir(parents=True, exist_ok=True)


# ======================================================
# 2. PROCESS DATE AND FILE NAMES
# ======================================================

process_date = pd.Timestamp("2026-01-31")
date_label = process_date.strftime("%Y%m%d")

raw_file = RAW_DIR / f"transactions_{date_label}.csv"
processed_file = PROCESSED_DIR / f"transactions_{date_label}.parquet"
template_file = TEMPLATES_DIR / "transactions_report_template.xlsx"
report_file = REPORTS_DIR / f"transactions_report_{date_label}.xlsx"


# ======================================================
# 3. CREATE SYNTHETIC RAW DATA
# ======================================================

df_raw_sample = pd.DataFrame(
    {
        "ID": ["T001", "T002", "T003", "T004", "T005"],
        "Client": ["Client A", "Client B", "Client A", "Client C", "Client B"],
        "Status": ["paid", "pending", "paid", "cancelled", "paid"],
        "Amount": ["1000,50", "2500,00", "1750,25", "500,00", "3200,75"],
        "Transaction Date": [
            "2026-01-03",
            "2026-01-10",
            "2026-01-15",
            "2026-01-20",
            "2026-01-28",
        ],
    }
)

export_csv(df_raw_sample, raw_file)


# ======================================================
# 4. LOAD RAW DATA
# ======================================================

df_raw = load_file(raw_file)

print("Raw data loaded:")
print(df_raw)


# ======================================================
# 5. CLEAN DATA
# ======================================================

df = clean_df(
    df_raw,
    cols_float=["amount"],
    cols_date=["transaction_date"],
)

print("\nClean data:")
print(df)


# ======================================================
# 6. VALIDATE DATA
# ======================================================

validate_not_empty(df, "transactions")

validate_required_columns(
    df,
    ["id", "client", "status", "amount", "transaction_date"],
)

validate_not_null(
    df,
    ["id", "client", "status", "amount"],
)

validate_unique(df, "id")

validate_non_negative(df, "amount")

validate_allowed_values(
    df,
    "status",
    ["paid", "pending", "cancelled"],
)

print("\nValidations passed.")


# ======================================================
# 7. EXPORT PROCESSED DATA
# ======================================================

export_parquet(df, processed_file)

print(f"\nProcessed data exported to: {processed_file}")


# ======================================================
# 8. BUILD REPORT DATA
# ======================================================

df_summary = (
    df.groupby("status", as_index=False)
    .agg(
        transactions=("id", "count"),
        amount=("amount", "sum"),
    )
)

total_amount = df["amount"].sum()

report_values = [
    "Synthetic Transactions Report",
    process_date,
    total_amount,
]

print("\nSummary table:")
print(df_summary)


# ======================================================
# 9. CREATE EXCEL TEMPLATE
# ======================================================

wb = Workbook()
ws = wb.active
ws.title = "Report"

ws["B2"] = "Transactions Report"
ws["B3"] = "Generated from synthetic data"

ws["B5"] = "Report name"
ws["B6"] = "Process date"
ws["B7"] = "Total amount"

ws["B10"] = "Summary"

wb.save(template_file)
wb.close()

print(f"\nTemplate created at: {template_file}")


# ======================================================
# 10. CREATE FINAL EXCEL REPORT
# ======================================================

create_report(
    template_path=template_file,
    output_path=report_file,
    items=[
        {
            "sheet": "Report",
            "cell": "C5",
            "data": report_values,
            "include_header": False,
        },
        {
            "sheet": "Report",
            "cell": "B11",
            "data": df_summary,
            "table_name": "SummaryTable",
            "number_cols": ["transactions", "amount"],
        },
    ],
)

print(f"\nExcel report created at: {report_file}")
print("\nEnd-to-end example completed successfully.")