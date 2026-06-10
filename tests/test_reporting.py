import pandas as pd
import pytest

pytest.importorskip("openpyxl")

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as ExcelImage

from dr_datakit.reporting import (
    copy_template,
    write_block,
    create_report,
)


def test_copy_template(tmp_path):
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "nested" / "report.xlsx"

    wb = Workbook()
    wb.save(template_path)
    wb.close()

    result_path = copy_template(template_path, output_path)

    assert result_path == output_path
    assert output_path.exists()


def test_write_block_scalar(tmp_path):
    output_path = tmp_path / "report.xlsx"

    wb = Workbook()
    ws = wb.active

    write_block(ws, "CODIGO", "C5")

    wb.save(output_path)
    wb.close()

    wb_loaded = load_workbook(output_path)
    ws_loaded = wb_loaded.active

    assert ws_loaded["C5"].value == "CODIGO"

    wb_loaded.close()


def test_write_block_dataframe_creates_table_by_default(tmp_path):
    output_path = tmp_path / "report.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    df = pd.DataFrame({
        "status": ["paid", "pending"],
        "amount": [100.5, 200.0],
        "rate": [0.8, 0.2],
    })

    write_block(
        ws,
        df,
        "B5",
        table_name="SummaryTable",
        number_cols=["amount"],
        percent_cols=["rate"],
    )

    wb.save(output_path)
    wb.close()

    wb_loaded = load_workbook(output_path)
    ws_loaded = wb_loaded["Report"]

    assert ws_loaded["B5"].value == "status"
    assert ws_loaded["C5"].value == "amount"
    assert ws_loaded["D5"].value == "rate"
    assert ws_loaded["B6"].value == "paid"
    assert ws_loaded["C6"].value == 100.5
    assert ws_loaded["D6"].value == 0.8
    assert "SummaryTable" in ws_loaded.tables

    wb_loaded.close()


def test_write_block_without_header_does_not_create_table(tmp_path):
    output_path = tmp_path / "report.xlsx"

    wb = Workbook()
    ws = wb.active

    df = pd.DataFrame({
        "value": ["CODIGO", pd.Timestamp("2026-01-31"), "CLP"],
    })

    write_block(
        ws,
        df,
        "C5",
        include_header=False,
    )

    wb.save(output_path)
    wb.close()

    wb_loaded = load_workbook(output_path)
    ws_loaded = wb_loaded.active

    assert ws_loaded["C5"].value == "CODIGO"
    assert ws_loaded["C6"].value == pd.Timestamp("2026-01-31").to_pydatetime()
    assert ws_loaded["C7"].value == "CLP"
    assert len(ws_loaded.tables) == 0

    wb_loaded.close()


def test_write_block_table_with_no_header_raises_error():
    wb = Workbook()
    ws = wb.active

    df = pd.DataFrame({
        "value": ["CODIGO", "CLP"],
    })

    with pytest.raises(ValueError):
        write_block(
            ws,
            df,
            "C5",
            include_header=False,
            as_table=True,
        )

    wb.close()


def test_write_block_transposed_with_row_formats(tmp_path):
    output_path = tmp_path / "report.xlsx"

    wb = Workbook()
    ws = wb.active

    df = pd.DataFrame({
        "-": ["Payment Rate", "Average Balance"],
        "jan-2026": [0.85, 1500.75],
        "feb-2026": [0.82, 1800.25],
    })

    write_block(
        ws,
        df,
        "B10",
        as_table=False,
        percent_rows=["Payment Rate"],
        number_rows=["Average Balance"],
    )

    wb.save(output_path)
    wb.close()

    wb_loaded = load_workbook(output_path)
    ws_loaded = wb_loaded.active

    assert ws_loaded["B10"].value == "-"
    assert ws_loaded["C10"].value == "jan-2026"
    assert ws_loaded["D10"].value == "feb-2026"
    assert ws_loaded["B11"].value == "Payment Rate"
    assert ws_loaded["C11"].value == 0.85
    assert ws_loaded["C11"].number_format == "0.00%"
    assert ws_loaded["B12"].value == "Average Balance"
    assert ws_loaded["C12"].value == 1500.75
    assert ws_loaded["C12"].number_format == "#,##0_);-#,##0;-"

    wb_loaded.close()


def test_create_report_from_template(tmp_path):
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "reports" / "report.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    ws["B2"] = "Template"
    wb.save(template_path)
    wb.close()

    df_params = pd.DataFrame({
        "value": ["CODIGO", pd.Timestamp("2026-01-31"), "CLP"],
    })

    df_summary = pd.DataFrame({
        "status": ["paid", "pending"],
        "amount": [100.5, 200.0],
    })

    result_path = create_report(
        template_path=template_path,
        output_path=output_path,
        items=[
            {
                "sheet": "Report",
                "cell": "C5",
                "data": df_params,
                "include_header": False,
            },
            {
                "sheet": "Report",
                "cell": "B10",
                "data": df_summary,
                "table_name": "ReportTable",
                "number_cols": ["amount"],
            },
        ],
    )

    assert result_path == output_path
    assert output_path.exists()

    wb_loaded = load_workbook(output_path)
    ws_loaded = wb_loaded["Report"]

    assert ws_loaded["B2"].value == "Template"
    assert ws_loaded["C5"].value == "CODIGO"
    assert ws_loaded["C6"].value == pd.Timestamp("2026-01-31").to_pydatetime()
    assert ws_loaded["C7"].value == "CLP"
    assert ws_loaded["B10"].value == "status"
    assert ws_loaded["C10"].value == "amount"
    assert "ReportTable" in ws_loaded.tables

    wb_loaded.close()


def test_create_report_preserves_template_images(tmp_path):
    pytest.importorskip("PIL")

    from PIL import Image as PILImage

    template_path = tmp_path / "template_with_image.xlsx"
    output_path = tmp_path / "reports" / "report_with_image.xlsx"
    image_path = tmp_path / "logo.png"

    img = PILImage.new("RGB", (20, 20), color="blue")
    img.save(image_path)

    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    ws.add_image(ExcelImage(str(image_path)), "B2")
    wb.save(template_path)
    wb.close()

    create_report(
        template_path=template_path,
        output_path=output_path,
        items=[
            {
                "sheet": "Report",
                "cell": "C5",
                "data": "CODIGO",
            },
        ],
    )

    wb_loaded = load_workbook(output_path)
    ws_loaded = wb_loaded["Report"]

    assert len(ws_loaded._images) == 1
    assert ws_loaded["C5"].value == "CODIGO"

    wb_loaded.close()