from pathlib import Path
from datetime import datetime, date
import re
import shutil

import pandas as pd

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils.cell import coordinate_from_string
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo


NUMBER_FORMAT = "#,##0_);-#,##0;-"
PERCENT_FORMAT = "0.00%"
DATE_FORMAT = "yyyy-mm-dd"
DEFAULT_TABLE_STYLE = "TableStyleMedium2"


def _ensure_xlsx_path(path) -> Path:
    """
    Ensure that a path has .xlsx extension.
    """
    path = Path(path)

    if path.suffix.lower() != ".xlsx":
        path = path.with_suffix(".xlsx")

    return path


def _ensure_parent_dir(path) -> Path:
    """
    Ensure that the parent directory of a file path exists.
    """
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    return path


def _parse_start_cell(start_cell: str) -> tuple[int, int]:
    """
    Parse an Excel cell reference into row and column index.
    """
    col_letter, row = coordinate_from_string(start_cell)
    col_idx = column_index_from_string(col_letter)

    return int(row), int(col_idx)


def _is_missing(value) -> bool:
    """
    Return True if a scalar value should be treated as missing.
    """
    if value is None:
        return True

    try:
        return bool(pd.isna(value))
    except (TypeError, ValueError):
        return False


def _clean_cell_value(value):
    """
    Convert pandas missing values to None for Excel compatibility.
    """
    if _is_missing(value):
        return None

    return value


def _sanitize_table_name(name: str) -> str:
    """
    Sanitize Excel table names.

    Excel table names cannot contain spaces or special characters
    and should not start with a number.
    """
    name = re.sub(r"[^A-Za-z0-9_]", "_", str(name))

    if not name:
        name = "Table"

    if name[0].isdigit():
        name = f"Table_{name}"

    return name[:255]


def _normalize_to_dataframe(data) -> tuple[pd.DataFrame, bool]:
    """
    Normalize input data to a DataFrame.

    Returns
    -------
    tuple[pd.DataFrame, bool]
        DataFrame and a boolean indicating whether the original input
        was a scalar value.
    """
    if isinstance(data, pd.DataFrame):
        return data.copy(), False

    if isinstance(data, pd.Series):
        return data.to_frame(), False

    if isinstance(data, (list, tuple)):
        return pd.DataFrame(data), False

    return pd.DataFrame([[data]]), True


def _prepare_output_dataframe(
    data,
    include_index: bool = False,
    transpose: bool = False,
) -> tuple[pd.DataFrame, bool]:
    """
    Prepare the DataFrame that will be written to Excel.
    """
    df, is_scalar = _normalize_to_dataframe(data)

    if transpose:
        df = df.T

    if include_index:
        index_name = df.index.name or "index"
        df = df.reset_index().rename(columns={"index": index_name})

    return df, is_scalar


def _apply_number_format(cell, value, value_type=None):
    """
    Apply default Excel number/date formats to a cell.
    """
    if value is None:
        return

    if value_type == "percent":
        cell.number_format = PERCENT_FORMAT
        return

    if value_type == "number":
        cell.number_format = NUMBER_FORMAT
        return

    if value_type == "date":
        cell.number_format = DATE_FORMAT
        return

    if isinstance(value, (datetime, pd.Timestamp, date)):
        cell.number_format = DATE_FORMAT

    elif isinstance(value, (int, float)):
        cell.number_format = NUMBER_FORMAT


def copy_template(template_path, output_path) -> Path:
    """
    Copy an Excel template to an output path.

    This is useful when the template contains logos, images,
    fixed text, styles, formulas or predefined sheet structure.
    """
    template_path = _ensure_xlsx_path(template_path)
    output_path = _ensure_xlsx_path(output_path)
    output_path = _ensure_parent_dir(output_path)

    if not template_path.exists():
        raise FileNotFoundError(f"Template not found: {template_path}")

    shutil.copyfile(template_path, output_path)

    return output_path


def write_block(
    ws,
    data,
    start_cell: str,
    include_header: bool = True,
    include_index: bool = False,
    transpose: bool = False,
    as_table: bool | None = None,
    table_name: str | None = None,
    number_cols=None,
    percent_cols=None,
    date_cols=None,
    number_rows=None,
    percent_rows=None,
    row_label_col: str | None = None,
    value_type: str | None = None,
    table_style: str = DEFAULT_TABLE_STYLE,
) -> None:
    """
    Write a block of data into an Excel worksheet.

    The input can be:
    - scalar value
    - pandas Series
    - pandas DataFrame
    - list or tuple

    Rules:
    - Scalars are written as 1x1 blocks.
    - DataFrames are written from start_cell.
    - If include_header=False, no Excel table is created.
    - If as_table=None, Excel tables are created only when headers are included.
    - Numeric values are displayed with thousands separator and no decimals.
    - Percentage values use 0.00%.
    - Date values use yyyy-mm-dd.
    """
    number_cols = set(number_cols or [])
    percent_cols = set(percent_cols or [])
    date_cols = set(date_cols or [])
    number_rows = set(number_rows or [])
    percent_rows = set(percent_rows or [])

    df, is_scalar = _prepare_output_dataframe(
        data=data,
        include_index=include_index,
        transpose=transpose,
    )

    if df.empty:
        return

    if is_scalar:
        include_header = False
        include_index = False
        as_table = False

    if not include_header and as_table is True:
        raise ValueError("Cannot create an Excel table when include_header=False.")

    if as_table is None:
        as_table = include_header and not is_scalar

    start_row, start_col = _parse_start_cell(start_cell)

    rows = list(
        dataframe_to_rows(
            df,
            index=False,
            header=include_header,
        )
    )

    if include_header and rows:
        rows[0] = [
            "" if value is None else str(value)
            for value in rows[0]
        ]

    output_col_labels = list(df.columns)

    if row_label_col is None and (number_rows or percent_rows):
        row_label_col = output_col_labels[0] if output_col_labels else None

    for row_idx, row_values in enumerate(rows, start=start_row):
        is_header_row = include_header and row_idx == start_row
        row_label = row_values[0] if row_values else None

        for col_idx, value in enumerate(row_values, start=start_col):
            value = _clean_cell_value(value)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            if is_header_row:
                continue

            col_pos = col_idx - start_col
            col_name = (
                output_col_labels[col_pos]
                if col_pos < len(output_col_labels)
                else None
            )

            is_row_label_cell = (
                row_label_col is not None
                and col_name == row_label_col
            )

            if col_name in percent_cols or row_label in percent_rows:
                if not is_row_label_cell:
                    _apply_number_format(cell, value, value_type="percent")

            elif col_name in number_cols or row_label in number_rows:
                if not is_row_label_cell:
                    _apply_number_format(cell, value, value_type="number")

            elif col_name in date_cols:
                _apply_number_format(cell, value, value_type="date")

            elif is_scalar and value_type is not None:
                _apply_number_format(cell, value, value_type=value_type)

            else:
                _apply_number_format(cell, value)

    if not as_table:
        return

    end_row = start_row + len(rows) - 1
    end_col = start_col + df.shape[1] - 1

    ref = (
        f"{get_column_letter(start_col)}{start_row}:"
        f"{get_column_letter(end_col)}{end_row}"
    )

    if table_name is None:
        table_name = f"Table_{ws.title}_{start_cell}"

    table_name = _sanitize_table_name(table_name)

    if table_name in ws.tables:
        del ws.tables[table_name]

    table = Table(displayName=table_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name=table_style,
        showRowStripes=True,
        showColumnStripes=False,
    )

    ws.add_table(table)


def create_report(
    template_path,
    output_path,
    items: list[dict] | None = None,
) -> Path:
    """
    Create an Excel report from a template and a list of blocks.

    Parameters
    ----------
    template_path:
        Path to the Excel template.

    output_path:
        Path where the report will be created.

    items:
        List of blocks to write.

        Example
        -------
        {
            "sheet": "Report",
            "cell": "B10",
            "data": df,
            "include_header": True,
            "as_table": True,
            "table_name": "ReportTable",
            "number_cols": ["amount"],
            "percent_cols": ["rate"],
        }
    """
    output_path = copy_template(template_path, output_path)

    wb = load_workbook(output_path)

    for item in items or []:
        sheet = item["sheet"]
        cell = item["cell"]
        data = item["data"]

        ws = wb[sheet] if sheet in wb.sheetnames else wb.create_sheet(sheet)

        write_block(
            ws=ws,
            data=data,
            start_cell=cell,
            include_header=item.get("include_header", True),
            include_index=item.get("include_index", False),
            transpose=item.get("transpose", False),
            as_table=item.get("as_table"),
            table_name=item.get("table_name"),
            number_cols=item.get("number_cols"),
            percent_cols=item.get("percent_cols"),
            date_cols=item.get("date_cols"),
            number_rows=item.get("number_rows"),
            percent_rows=item.get("percent_rows"),
            row_label_col=item.get("row_label_col"),
            value_type=item.get("value_type"),
            table_style=item.get("table_style", DEFAULT_TABLE_STYLE),
        )

    wb.save(output_path)
    wb.close()

    return output_path